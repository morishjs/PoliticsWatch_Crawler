#-*- coding: utf-8 -*-
from openpyxl import Workbook
from web_crawler import spider
import json


class e:
    def __init__(self):
        self.wb = Workbook()
        self.sheet = self.wb.active
        self.sheet['A1'] = 'name_kr'
        self.sheet['B1'] = 'r_count'
        self.sheet['C1'] = 'r_percentage'
        self.sheet['D1'] = 'party'
        self.sheet['E1'] = 'district'
        self.sheet['F1'] = 'homepage'
        self.sheet['G1'] = 'a_count'
        self.sheet['G2'] = 20
        self.isFirst = True
        self.pLen = 0

    #출석 횟수 구하기
    def cal_c(self,l):

        all_p, attend, leave, absent = spider(l)

        if(self.isFirst):
            self.pLen = len(all_p)
            # 모든 이름 넣기(초기화)
            for i in range(0, len(all_p)):
                d = self.sheet.cell(row=i + 2, column=1)
                count = self.sheet.cell(row=i + 2, column=2)
                d.value = all_p[i]
                count.value = 0



            self.isFirst = False


        for rowNum in range(2, self.sheet.max_row+1):
            d = self.sheet.cell(row=rowNum, column=1)
            if d.value in attend:
                count = self.sheet.cell(row=rowNum, column=2)
                count.value += 1

        self.wb.save('test.xlsx')

    #출석률 구하기
    def cal_r(self):
        # 총 본회의
        a = self.sheet['G2']
        for i in range(0,self.pLen):
            #average
            r = self.sheet.cell(row=i + 2, column=3)
            #출석횟수
            c = self.sheet.cell(row=i + 2, column=2)
            r.value = float(c.value) / float(a.value) * 100
        self.wb.save('test.xlsx')


    #기본정보 엑셀에 입력하기
    def dump_info(self):
        s = open("polist.json").read()
        dict = json.loads(s)


        # print len(dict)
        for person in dict:
            i = 2
            while(i <= self.sheet.max_row):
                name = self.sheet.cell(row=i, column=1)
                if(person["name_kr"] == name.value):
                    party = self.sheet.cell(row=i, column=4)
                    district = self.sheet.cell(row=i, column=5)
                    homepage = self.sheet.cell(row=i, column=6)

                    party.value = person["party"]
                    district.value = person["district"]
                    homepage.value = person["homepage"]

                    break
                i += 1

        self.wb.save('test.xlsx')



